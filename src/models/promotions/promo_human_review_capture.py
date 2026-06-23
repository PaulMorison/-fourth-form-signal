from __future__ import annotations

"""Phase 5W/5X — human review capture, filled ingestion, and buyer learning pack."""

from datetime import datetime, timezone
from importlib.util import find_spec
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from models.promotions.promo_shadow_observation_journal import (
    HUMAN_BUYER_DECISIONS,
    MERGE_KEY_COLUMNS,
    PHASE5U_DIAGNOSTICS_DIR,
)

DEFAULT_DIAGNOSTICS_DIR = Path("Diagnostics/phase5w01_human_review_capture")
PHASE5X_DIAGNOSTICS_DIR = Path("Diagnostics/phase5x01_filled_human_review_learning")
PHASE5T_JOURNAL_PATH = Path("Diagnostics/phase5t01_shadow_observation_journal/SHADOW_TOP_100_OBSERVATION_JOURNAL.csv")
PHASE5U_SCORED_PATH = PHASE5U_DIAGNOSTICS_DIR / "phase5u01_shadow_scored_outcomes.csv"
PHASE5V_WEIGHTED_PATH = Path("Diagnostics/phase5v01_lesson_weighted_updates/phase5v01_lesson_weighted_training_frame.csv")
WORKBOOK_FILENAME = "SHADOW_TOP_100_BUYER_REVIEW_WORKBOOK.xlsx"
FILLED_REVIEW_FILENAME = "SHADOW_TOP_100_HUMAN_REVIEW_FILLED.csv"
FILLED_REVIEW_PATH = DEFAULT_DIAGNOSTICS_DIR / FILLED_REVIEW_FILENAME
STATUS_WORKBOOK_FILENAME = "SHADOW_TOP_100_BUYER_REVIEW_STATUS.xlsx"

HUMAN_OVERRIDE_REASONS = (
    "BRAIN_TOO_AGGRESSIVE",
    "BRAIN_TOO_CONSERVATIVE",
    "GOVERNANCE_TOO_AGGRESSIVE",
    "GOVERNANCE_TOO_CONSERVATIVE",
    "LOW_CONFIDENCE_STOCK_DATA",
    "SUPPLIER_CONSTRAINT",
    "KNOWN_LOCAL_DEMAND",
    "RANGE_TRUST_PROTECTION",
    "LONG_TAIL_BASKET_PROTECTION",
    "PROMO_NOT_COMPELLING",
    "OVERSTOCK_RUN_DOWN",
    "OTHER",
)

FILLABLE_COLUMNS = (
    "human_buyer_decision",
    "human_order_units",
    "human_decision_reason",
    "human_override_flag",
    "human_override_reason",
    "human_confidence_score",
    "human_reviewer",
    "human_reviewed_at",
    "human_notes",
)

TOP_100_REVIEW_COLUMNS = (
    "shadow_run_id",
    "store_number",
    "promotion_id",
    "promotion_name",
    "promotion_start_date",
    "promotion_end_date",
    "sku_number",
    "sku_description",
    "department",
    "category",
    "brain_validated_action_label",
    "brain_validated_expected_value",
    "final_governed_action_label",
    "final_governed_order_units",
    "shadow_candidate_class",
    "shadow_candidate_score",
    "shadow_candidate_rank",
    "validated_alpha_pattern_label",
    "expected_shadow_learning_question",
    "current_soh",
    "expected_soh_at_promo_start_before_order",
    "optimal_base_soh_units",
    "target_day_one_promo_soh",
    "target_end_promo_soh",
    "expected_promo_uplift_units",
    "economic_net_value_score",
    "shadow_expected_learning_value",
    "mission_sku_score",
    "long_tail_sku_flag",
    "basket_attachment_source_quality",
    "segment_historical_bias_pct",
    *FILLABLE_COLUMNS,
)

FEEDBACK_COLUMNS = (
    "human_feedback_signal",
    "human_feedback_weight",
    "human_disagrees_with_brain_flag",
    "human_disagrees_with_governed_flag",
    "human_accepts_brain_flag",
    "human_accepts_governed_flag",
    "human_override_direction",
    "human_feedback_learning_note",
)

TRAINING_SIGNAL_COLUMNS = (
    "human_feedback_quality",
    "human_confidence_weighted_signal",
    "human_agrees_with_brain_flag",
    "human_agrees_with_governed_flag",
    "human_overrides_both_flag",
    "human_quantity_delta_vs_brain",
    "human_quantity_delta_vs_governed",
    "human_feedback_training_ready_flag",
)

BUYER_LEARNING_THEMES = (
    "BUYER_ACCEPTS_BRAIN",
    "BUYER_ACCEPTS_GOVERNED",
    "BUYER_OVERRIDES_BOTH",
    "BRAIN_TOO_AGGRESSIVE",
    "BRAIN_TOO_CONSERVATIVE",
    "GOVERNANCE_TOO_AGGRESSIVE",
    "GOVERNANCE_TOO_CONSERVATIVE",
    "LONG_TAIL_BASKET_PROTECTION",
    "SUPPLIER_CONSTRAINT",
    "DATA_QUALITY_BLOCK",
    "KNOWN_LOCAL_DEMAND",
)

VALIDATION_COLUMNS = (
    "human_review_valid_flag",
    "human_review_status",
    "human_review_validation_error",
    "human_review_merge_status",
    "human_decision_merge_status",
)

INSTRUCTIONS_MARKDOWN = """# Shadow Top 100 Buyer Review — Instructions

This workbook is for internal shadow observation only. Your decisions do not place orders or change production recommendations.

## How to review
1. Start on **Top_100_Review** — one row per shadow SKU.
2. Read the system recommendation columns first (brain vs governed).
3. Fill only the yellow **human review** columns on the right.
4. Use **Allowed_Values** for valid decision and override-reason codes.
5. Save completed rows and export to SHADOW_TOP_100_HUMAN_REVIEW_FILLED.csv for ingestion.

## Fillable fields
- **human_buyer_decision** — what you would do (required when reviewing a row).
- **human_order_units** — order quantity if buying (0 allowed for hold/run-down).
- **human_override_flag** — YES if you are overriding the governed recommendation.
- **human_override_reason** — why you disagree (see Allowed_Values).
- **human_confidence_score** — 0 to 100.
- **human_reviewer** — your name or buyer code (required when decision is filled).
- **human_reviewed_at** — date reviewed (YYYY-MM-DD).
- **human_notes** — optional free text.

## Important
- segment_historical_bias_pct is shown as a percentage (e.g. 17.9 means 17.9%).
- Leave rows blank if not yet reviewed — they remain pending.
- Do not edit identity or system recommendation columns.
"""


def _excel_supported() -> bool:
    return find_spec("openpyxl") is not None


def _numeric(series: pd.Series, default: float = 0.0) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(default)


def _round_frame(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    for col in out.columns:
        if pd.api.types.is_numeric_dtype(out[col]):
            out[col] = _numeric(out[col]).round(4)
    return out


def _select_review_columns(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    if "human_notes" not in out.columns:
        out["human_notes"] = ""
    for col in FILLABLE_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    cols = [c for c in TOP_100_REVIEW_COLUMNS if c in out.columns]
    return _round_frame(out[cols])


def _enrich_source_frame(
    journal_df: pd.DataFrame,
    scored_df: pd.DataFrame | None = None,
    lesson_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    frame = journal_df.copy()
    merge_cols = [c for c in ("store_number", "promotion_id", "sku_number") if c in frame.columns]

    for extra, label in ((scored_df, "scored"), (lesson_df, "lesson")):
        if extra is None or extra.empty or not merge_cols:
            continue
        add_cols = [c for c in extra.columns if c not in frame.columns or c in merge_cols]
        right = extra[add_cols].drop_duplicates(subset=merge_cols, keep="first").copy()
        for col in merge_cols:
            frame[col] = frame[col].astype(str)
            right[col] = right[col].astype(str)
        frame = frame.merge(right, on=merge_cols, how="left", suffixes=("", f"_{label}"))

    return frame


def build_allowed_values_sheet() -> pd.DataFrame:
    rows = [{"field_name": "human_buyer_decision", "allowed_value": v, "description": "Buyer action for this shadow SKU"} for v in HUMAN_BUYER_DECISIONS]
    rows += [{"field_name": "human_override_reason", "allowed_value": v, "description": "Why buyer overrode system recommendation"} for v in HUMAN_OVERRIDE_REASONS]
    rows.append({"field_name": "human_override_flag", "allowed_value": "YES", "description": "Buyer is overriding governed recommendation"})
    rows.append({"field_name": "human_override_flag", "allowed_value": "NO", "description": "Buyer aligns with governed recommendation"})
    rows.append({"field_name": "human_confidence_score", "allowed_value": "0-100", "description": "Buyer confidence percentage"})
    return pd.DataFrame(rows)


def build_instructions_sheet() -> pd.DataFrame:
    return pd.DataFrame([{"section": "instructions", "content": line} for line in INSTRUCTIONS_MARKDOWN.splitlines() if line.strip()])


def build_review_summary_sheet(frame: pd.DataFrame, analytics: pd.DataFrame | None = None) -> pd.DataFrame:
    rows = [
        {"metric": "total_review_rows", "value": int(len(frame))},
        {"metric": "shadow_run_id", "value": str(frame["shadow_run_id"].iloc[0]) if len(frame) and "shadow_run_id" in frame.columns else ""},
        {"metric": "workbook_generated_at", "value": datetime.now(timezone.utc).isoformat()},
        {"metric": "release_recommendation", "value": "NO_RELEASE"},
        {"metric": "primary_blocker", "value": "model_bias_dangerously_negative"},
        {"metric": "operational_trial", "value": "NOT_APPROVED"},
    ]
    if analytics is not None and not analytics.empty:
        for col in analytics.columns:
            rows.append({"metric": col, "value": str(analytics.iloc[0][col])})
    return pd.DataFrame(rows)


def build_human_review_workbook(
    journal_df: pd.DataFrame,
    *,
    scored_df: pd.DataFrame | None = None,
    lesson_df: pd.DataFrame | None = None,
    analytics: pd.DataFrame | None = None,
    output_path: Path | None = None,
) -> dict[str, Any]:
    """Build buyer-friendly shadow review workbook and sheet frames."""
    enriched = _enrich_source_frame(journal_df, scored_df=scored_df, lesson_df=lesson_df)
    top_100 = _select_review_columns(enriched)

    long_tail = top_100[
        top_100.get("long_tail_sku_flag", pd.Series("NO", index=top_100.index)).astype(str).eq("YES")
        | _numeric(top_100.get("mission_sku_score", pd.Series(0, index=top_100.index))).ge(45)
    ].copy()

    gov_conservative = enriched.loc[
        enriched.get("lesson_learned_label", pd.Series("", index=enriched.index)).astype(str).eq("GOVERNANCE_TOO_CONSERVATIVE")
    ].copy()
    if gov_conservative.empty:
        gov_conservative = enriched.head(0).copy()
    gov_conservative = _select_review_columns(gov_conservative) if not gov_conservative.empty else top_100.head(0)

    brain_mismatch = enriched.loc[
        enriched.get("brain_validated_action_label", pd.Series("", index=enriched.index)).astype(str)
        != enriched.get("final_governed_action_label", pd.Series("", index=enriched.index)).astype(str)
    ].copy()
    brain_mismatch = _select_review_columns(brain_mismatch)

    allowed = build_allowed_values_sheet()
    instructions = build_instructions_sheet()
    summary = build_review_summary_sheet(top_100, analytics=analytics)

    sheets = {
        "Top_100_Review": top_100,
        "Long_Tail_Mission_SKUs": _select_review_columns(long_tail) if not long_tail.empty else top_100.head(0),
        "Governance_Too_Conservative": gov_conservative,
        "Brain_vs_Governed_Mismatch": brain_mismatch,
        "Allowed_Values": allowed,
        "Instructions": instructions,
        "Review_Summary": summary,
    }

    workbook_path = output_path
    generated = False
    if workbook_path is not None and _excel_supported():
        workbook_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
            for name, sheet_df in sheets.items():
                sheet_df.to_excel(writer, sheet_name=name, index=False)
        _style_workbook(workbook_path, fillable_columns=list(FILLABLE_COLUMNS))
        generated = workbook_path.exists()

    return {
        "sheets": sheets,
        "workbook_path": str(workbook_path) if workbook_path and generated else None,
        "workbook_generated": generated,
        "review_rows": int(len(top_100)),
    }


def _style_workbook(path: Path, *, fillable_columns: list[str]) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    fillable_fill = PatternFill("solid", fgColor="FFF2CC")
    readonly_fill = PatternFill("solid", fgColor="F2F2F2")
    bold = Font(bold=True)

    for sheet_name in wb.sheetnames:
        if sheet_name == "Instructions":
            continue
        ws = wb[sheet_name]
        if ws.max_row < 1 or ws.max_column < 1:
            continue
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        headers = [cell.value for cell in ws[1]]
        fillable_idx = {headers.index(c) + 1 for c in fillable_columns if c in headers}
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = bold
            cell.fill = fillable_fill if col_idx in fillable_idx else header_fill
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = (
                    fillable_fill if col_idx in fillable_idx else readonly_fill
                )

    if "Instructions" in wb.sheetnames:
        wb["Instructions"].column_dimensions["A"].width = 18
        wb["Instructions"].column_dimensions["B"].width = 100

    wb.save(path)


def _parse_review_date(value: Any) -> bool:
    if value is None or str(value).strip() == "":
        return True
    try:
        pd.to_datetime(value)
        return True
    except (ValueError, TypeError):
        return False


def validate_human_review_input(human_df: pd.DataFrame) -> pd.DataFrame:
    """Validate filled human review rows."""
    if human_df.empty:
        return human_df.copy()

    out = human_df.copy()
    for col in MERGE_KEY_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    dup_keys = out.duplicated(subset=list(MERGE_KEY_COLUMNS), keep=False)

    statuses, valid_flags, errors, merge_status = [], [], [], []
    for idx, row in out.iterrows():
        if bool(dup_keys.loc[idx]):
            statuses.append("INVALID")
            valid_flags.append("NO")
            errors.append("duplicate_review_row")
            merge_status.append("MERGE_REJECTED")
            continue

        decision = str(row.get("human_buyer_decision", "")).strip()
        if not decision:
            statuses.append("PENDING")
            valid_flags.append("NO")
            errors.append("")
            merge_status.append("NOT_MERGED")
            continue

        row_errors: list[str] = []
        if decision not in HUMAN_BUYER_DECISIONS:
            row_errors.append("invalid_human_buyer_decision")
        override_reason = str(row.get("human_override_reason", "")).strip()
        if override_reason and override_reason not in HUMAN_OVERRIDE_REASONS:
            row_errors.append("invalid_human_override_reason")
        if float(_numeric(pd.Series([row.get("human_order_units", 0)])).iloc[0]) < 0:
            row_errors.append("negative_human_order_units")
        conf = pd.to_numeric(row.get("human_confidence_score", np.nan), errors="coerce")
        if not pd.isna(conf) and (float(conf) < 0 or float(conf) > 100):
            row_errors.append("human_confidence_out_of_range")
        if not str(row.get("human_reviewer", "")).strip():
            row_errors.append("missing_human_reviewer")
        if not _parse_review_date(row.get("human_reviewed_at")):
            row_errors.append("invalid_human_reviewed_at")

        if row_errors:
            statuses.append("INVALID")
            valid_flags.append("NO")
            errors.append(";".join(row_errors))
            merge_status.append("MERGE_REJECTED")
        else:
            statuses.append("COMPLETE")
            valid_flags.append("YES")
            errors.append("")
            merge_status.append("MERGED")

    out["human_review_valid_flag"] = valid_flags
    out["human_review_status"] = statuses
    out["human_review_validation_error"] = errors
    out["human_review_merge_status"] = merge_status
    out["human_decision_merge_status"] = merge_status
    return out


def load_filled_human_review_file(path: Path | str | None = None) -> tuple[pd.DataFrame | None, bool]:
    """Load filled human review CSV if present."""
    review_path = Path(path) if path is not None else FILLED_REVIEW_PATH
    if not review_path.exists():
        return None, False
    frame = pd.read_csv(review_path)
    for col in MERGE_KEY_COLUMNS:
        if col not in frame.columns:
            raise ValueError(f"Missing required identity column: {col}")
    return frame, True


def validate_filled_human_review_file(human_df: pd.DataFrame | None) -> pd.DataFrame:
    """Validate filled human review file rows."""
    if human_df is None or human_df.empty:
        return pd.DataFrame(columns=list(MERGE_KEY_COLUMNS) + list(FILLABLE_COLUMNS) + list(VALIDATION_COLUMNS))
    return validate_human_review_input(human_df)


def _derive_feedback_signals(row: pd.Series) -> dict[str, Any]:
    decision = str(row.get("human_buyer_decision", "")).strip()
    if not decision or str(row.get("human_review_status", "")) != "COMPLETE":
        return {
            "human_feedback_signal": "PENDING",
            "human_feedback_weight": 0.0,
            "human_disagrees_with_brain_flag": "NO",
            "human_disagrees_with_governed_flag": "NO",
            "human_accepts_brain_flag": "NO",
            "human_accepts_governed_flag": "NO",
            "human_override_direction": "",
            "human_feedback_learning_note": "Awaiting buyer review.",
        }

    brain = str(row.get("brain_validated_action_label", "")).strip()
    governed = str(row.get("final_governed_action_label", "")).strip()
    override_reason = str(row.get("human_override_reason", "")).strip()
    override_flag = str(row.get("human_override_flag", "")).strip().upper() == "YES"

    accepts_brain = decision == "BUY_AS_BRAIN_SUGGESTED"
    accepts_governed = decision == "BUY_AS_GOVERNED_SUGGESTED"
    disagrees_brain = accepts_governed or decision in {"NO_BUY_RUN_DOWN", "HOLD", "BUY_DIFFERENT_QUANTITY", "BLOCKED_DATA_QUALITY"} or override_flag
    disagrees_governed = accepts_brain or decision in {"BUY_DIFFERENT_QUANTITY", "NO_BUY_RUN_DOWN", "BLOCKED_DATA_QUALITY"} or override_flag

    signal = "HUMAN_ALIGNED_WITH_GOVERNED"
    note = "Human decision recorded for shadow learning."
    weight = 1.0

    if decision == "BUY_AS_BRAIN_SUGGESTED" and brain != governed:
        signal = "ACCEPT_BRAIN_OVER_GOVERNED"
        note = "Human accepted brain over governed; reinforce brain pattern with review."
    elif decision == "BUY_AS_GOVERNED_SUGGESTED":
        signal = "ACCEPT_GOVERNED_OVER_BRAIN"
        note = "Human accepted governed over brain; reinforce governance."
    elif decision == "BUY_DIFFERENT_QUANTITY":
        signal = "QUANTITY_CALIBRATION"
        note = "Human chose different quantity; create quantity calibration lesson."
    elif decision == "BLOCKED_DATA_QUALITY":
        signal = "DATA_QUALITY_BLOCK"
        note = "Human blocked on data quality; prioritise data repair."
        weight = 0.5
    elif decision == "SUPPLIER_UNAVAILABLE":
        signal = "SUPPLIER_CONSTRAINT"
        note = "Supplier issue cited; not a model failure signal."
        weight = 0.5
    elif override_reason == "LONG_TAIL_BASKET_PROTECTION" or str(row.get("long_tail_sku_flag", "")) == "YES" and override_flag:
        signal = "LONG_TAIL_PROTECTION"
        note = "Human cited long-tail basket protection; reinforce long-tail features."
    elif override_reason in {"BRAIN_TOO_AGGRESSIVE", "BRAIN_TOO_CONSERVATIVE"}:
        signal = "BRAIN_CALIBRATION"
        note = "Human disagreed with brain aggressiveness; review action classifier."

    direction = override_reason or ("NONE" if not override_flag else "OTHER")

    return {
        "human_feedback_signal": signal,
        "human_feedback_weight": weight,
        "human_disagrees_with_brain_flag": "YES" if disagrees_brain else "NO",
        "human_disagrees_with_governed_flag": "YES" if disagrees_governed else "NO",
        "human_accepts_brain_flag": "YES" if accepts_brain else "NO",
        "human_accepts_governed_flag": "YES" if accepts_governed else "NO",
        "human_override_direction": direction,
        "human_feedback_learning_note": note,
    }


def _derive_training_signals(row: pd.Series) -> dict[str, Any]:
    base = _derive_feedback_signals(row)
    decision = str(row.get("human_buyer_decision", "")).strip()
    status = str(row.get("human_review_status", ""))
    if not decision or status != "COMPLETE":
        return {
            **base,
            "human_feedback_quality": "PENDING",
            "human_confidence_weighted_signal": 0.0,
            "human_agrees_with_brain_flag": "NO",
            "human_agrees_with_governed_flag": "NO",
            "human_overrides_both_flag": "NO",
            "human_quantity_delta_vs_brain": 0.0,
            "human_quantity_delta_vs_governed": 0.0,
            "human_feedback_training_ready_flag": "NO",
        }

    conf = float(_numeric(pd.Series([row.get("human_confidence_score", 0)])).iloc[0])
    human_units = float(_numeric(pd.Series([row.get("human_order_units", 0)])).iloc[0])
    gov_units = float(_numeric(pd.Series([row.get("final_governed_order_units", 0)])).iloc[0])
    override_flag = str(row.get("human_override_flag", "")).strip().upper() == "YES"
    override_reason = str(row.get("human_override_reason", "")).strip()
    signal = base["human_feedback_signal"]

    agrees_brain = decision == "BUY_AS_BRAIN_SUGGESTED"
    agrees_gov = decision == "BUY_AS_GOVERNED_SUGGESTED"
    overrides_both = (
        decision in {"BUY_DIFFERENT_QUANTITY", "NO_BUY_RUN_DOWN", "HOLD", "BLOCKED_DATA_QUALITY", "SUPPLIER_UNAVAILABLE", "OTHER"}
        or (override_flag and not agrees_brain and not agrees_gov)
    )

    delta_gov = round(human_units - gov_units, 4)
    delta_brain = 0.0 if agrees_brain else delta_gov

    if conf >= 70:
        quality = "HIGH"
    elif conf >= 50:
        quality = "MEDIUM"
    else:
        quality = "LOW"

    weight = float(base["human_feedback_weight"])
    if quality == "HIGH":
        weight *= 1.0 + (conf / 100.0) * 0.5
    elif quality == "LOW":
        weight *= 0.5

    training_ready = (
        status == "COMPLETE"
        and str(row.get("human_review_valid_flag", "")) == "YES"
        and conf >= 50
        and signal not in {"DATA_QUALITY_BLOCK"}
        and decision != "BLOCKED_DATA_QUALITY"
    )

    note = base["human_feedback_learning_note"]
    if decision == "BLOCKED_DATA_QUALITY":
        note = "Train data-quality system; do not treat as demand-brain failure."
        training_ready = False
    elif decision == "SUPPLIER_UNAVAILABLE":
        note = "Train supplier reliability; not a brain failure signal."
        training_ready = conf >= 50
    elif overrides_both:
        note = "Human overrode both paths; await actual outcomes before declaring human right."

    return {
        **base,
        "human_feedback_quality": quality,
        "human_confidence_weighted_signal": round(weight * (conf / 100.0), 4),
        "human_agrees_with_brain_flag": "YES" if agrees_brain else "NO",
        "human_agrees_with_governed_flag": "YES" if agrees_gov else "NO",
        "human_overrides_both_flag": "YES" if overrides_both else "NO",
        "human_quantity_delta_vs_brain": delta_brain,
        "human_quantity_delta_vs_governed": delta_gov,
        "human_feedback_training_ready_flag": "YES" if training_ready else "NO",
        "human_feedback_learning_note": note,
        "human_feedback_weight": round(weight, 4),
    }


def merge_human_review_feedback(
    journal_df: pd.DataFrame,
    human_df: pd.DataFrame | None = None,
    *,
    scored_df: pd.DataFrame | None = None,
    lesson_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Merge validated human feedback into shadow journal with learning signals."""
    base = _enrich_source_frame(journal_df, scored_df=scored_df, lesson_df=lesson_df)
    if human_df is None or human_df.empty:
        validated = pd.DataFrame(columns=list(MERGE_KEY_COLUMNS) + list(FILLABLE_COLUMNS) + list(VALIDATION_COLUMNS))
    else:
        validated = validate_human_review_input(human_df)

    merge_cols = list(MERGE_KEY_COLUMNS)
    human_cols = [c for c in FILLABLE_COLUMNS if c in validated.columns]
    status_cols = list(VALIDATION_COLUMNS)

    if validated.empty:
        merged = base.copy()
        for col in FILLABLE_COLUMNS:
            if col not in merged.columns:
                merged[col] = ""
        for col in VALIDATION_COLUMNS:
            merged[col] = "PENDING" if col == "human_review_status" else ("NOT_MERGED" if col in {"human_review_merge_status", "human_decision_merge_status"} else "")
        if "human_review_valid_flag" in merged.columns:
            merged["human_review_valid_flag"] = "NO"
    else:
        left = base.copy()
        right = validated[merge_cols + human_cols + status_cols].copy()
        for col in merge_cols:
            left[col] = left[col].astype(str)
            right[col] = right[col].astype(str)
        merged = left.merge(right, on=merge_cols, how="left", suffixes=("_journal", ""))
        for col in human_cols + status_cols:
            journal_col = f"{col}_journal"
            if journal_col in merged.columns and col in merged.columns:
                empty = merged[col].isna() | merged[col].astype(str).str.strip().eq("")
                merged[col] = merged[col].where(~empty, merged[journal_col])
                merged = merged.drop(columns=[journal_col])
        for col in FILLABLE_COLUMNS:
            if col not in merged.columns:
                merged[col] = ""
        for col in VALIDATION_COLUMNS:
            if col not in merged.columns:
                merged[col] = np.where(
                    merged.get("human_buyer_decision", pd.Series("", index=merged.index)).astype(str).str.strip().eq(""),
                    "PENDING" if col == "human_review_status" else ("NOT_MERGED" if col == "human_review_merge_status" else "NO"),
                    "",
                )

    feedback_rows = []
    for _, row in merged.iterrows():
        feedback_rows.append(_derive_training_signals(row))
    feedback = pd.DataFrame(feedback_rows, index=merged.index)
    for col in FEEDBACK_COLUMNS + TRAINING_SIGNAL_COLUMNS:
        if col in feedback.columns:
            merged[col] = feedback[col]

    decision_empty = merged.get("human_buyer_decision", pd.Series("", index=merged.index)).astype(str).str.strip().eq("")
    status = merged.get("human_review_status", pd.Series("", index=merged.index)).astype(str)
    pending_mask = decision_empty | status.isin({"", "nan", "None"})
    merged.loc[pending_mask, "human_review_status"] = "PENDING"
    merged.loc[pending_mask, "human_review_valid_flag"] = "NO"
    merged.loc[pending_mask, "human_review_merge_status"] = "NOT_MERGED"
    merged.loc[pending_mask, "human_decision_merge_status"] = "NOT_MERGED"
    merged.loc[pending_mask, "human_review_validation_error"] = ""
    if "human_decision_merge_status" not in merged.columns:
        merged["human_decision_merge_status"] = merged.get("human_review_merge_status", "NOT_MERGED")
    return merged


def build_override_analytics(merged_frame: pd.DataFrame) -> pd.DataFrame:
    """Summarise human override patterns from merged review frame."""
    if merged_frame.empty:
        return pd.DataFrame([{"review_completion_rate": 0.0}])

    status = merged_frame.get("human_review_status", pd.Series("PENDING", index=merged_frame.index)).astype(str)
    decision = merged_frame.get("human_buyer_decision", pd.Series("", index=merged_frame.index)).astype(str)
    complete = status.eq("COMPLETE")
    override_flag = merged_frame.get("human_override_flag", pd.Series("", index=merged_frame.index)).astype(str).str.upper().eq("YES")
    override_reason = merged_frame.get("human_override_reason", pd.Series("", index=merged_frame.index)).astype(str)
    learning_value = _numeric(merged_frame.get("shadow_expected_learning_value", pd.Series(0, index=merged_frame.index)))

    reason_counts = override_reason[override_reason.ne("")].value_counts()
    reason_by_value = (
        merged_frame.loc[override_reason.ne("")]
        .groupby("human_override_reason")["shadow_expected_learning_value"]
        .sum()
        .sort_values(ascending=False)
    )

    long_tail_overrides = int(
        (override_flag & merged_frame.get("long_tail_sku_flag", pd.Series("NO", index=merged_frame.index)).astype(str).eq("YES")).sum()
    )
    gov_conservative_overrides = int(override_reason.eq("GOVERNANCE_TOO_CONSERVATIVE").sum())
    brain_aggressive_overrides = int(override_reason.eq("BRAIN_TOO_AGGRESSIVE").sum())
    brain_conservative_overrides = int(override_reason.eq("BRAIN_TOO_CONSERVATIVE").sum())
    dq_blocks = int(decision.eq("BLOCKED_DATA_QUALITY").sum())

    return pd.DataFrame([{
        "review_completion_rate": round(float(complete.mean() * 100.0), 2),
        "completed_reviews": int(complete.sum()),
        "pending_reviews": int(status.eq("PENDING").sum()),
        "accepted_brain_count": int(decision.eq("BUY_AS_BRAIN_SUGGESTED").sum()),
        "accepted_governed_count": int(decision.eq("BUY_AS_GOVERNED_SUGGESTED").sum()),
        "different_quantity_count": int(decision.eq("BUY_DIFFERENT_QUANTITY").sum()),
        "no_buy_run_down_count": int(decision.eq("NO_BUY_RUN_DOWN").sum()),
        "blocked_data_quality_count": dq_blocks,
        "supplier_unavailable_count": int(decision.eq("SUPPLIER_UNAVAILABLE").sum()),
        "override_count": int(override_flag.sum()),
        "average_human_confidence": round(float(_numeric(merged_frame.get("human_confidence_score", pd.Series(0, index=merged_frame.index))).mean()), 2),
        "override_reasons_by_count": ";".join(f"{k}:{int(v)}" for k, v in reason_counts.items()),
        "override_reasons_by_learning_value": ";".join(f"{k}:{round(float(v), 2)}" for k, v in reason_by_value.items()),
        "long_tail_overrides": long_tail_overrides,
        "governance_conservative_overrides": gov_conservative_overrides,
        "brain_too_aggressive_overrides": brain_aggressive_overrides,
        "brain_too_conservative_overrides": brain_conservative_overrides,
        "data_quality_blocks": dq_blocks,
    }])


def build_phase5w_release_gate(merged_frame: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame([{
        "recommendation": "NO_RELEASE",
        "shadow_recommendation": "SHADOW_TOP_100_REVIEW",
        "review_rows": int(len(merged_frame)),
        "auto_order_created": "NO",
        "governed_actions_overwritten": "NO",
        "production_predictions_overwritten": "NO",
        "primary_blocker": "model_bias_dangerously_negative",
        "reason": "Human review capture is internal only; customer release not earned.",
    }])


def write_phase5w_diagnostics(
    *,
    journal_df: pd.DataFrame | None = None,
    human_df: pd.DataFrame | None = None,
    scored_df: pd.DataFrame | None = None,
    lesson_df: pd.DataFrame | None = None,
    diagnostics_dir: Path = DEFAULT_DIAGNOSTICS_DIR,
) -> dict[str, Any]:
    diagnostics_dir.mkdir(parents=True, exist_ok=True)

    if journal_df is None:
        if not PHASE5T_JOURNAL_PATH.exists():
            raise FileNotFoundError(f"Shadow journal not found: {PHASE5T_JOURNAL_PATH}")
        journal_df = pd.read_csv(PHASE5T_JOURNAL_PATH)
    if scored_df is None and PHASE5U_SCORED_PATH.exists():
        scored_df = pd.read_csv(PHASE5U_SCORED_PATH)
    if lesson_df is None and PHASE5V_WEIGHTED_PATH.exists():
        lesson_df = pd.read_csv(PHASE5V_WEIGHTED_PATH)

    filled_path = diagnostics_dir / FILLED_REVIEW_FILENAME
    if human_df is None and filled_path.exists():
        human_df = pd.read_csv(filled_path)

    merged = merge_human_review_feedback(journal_df, human_df, scored_df=scored_df, lesson_df=lesson_df)
    validation = validate_human_review_input(human_df) if human_df is not None and not human_df.empty else pd.DataFrame()
    if validation.empty:
        validation = merged[list(MERGE_KEY_COLUMNS) + list(VALIDATION_COLUMNS)].copy() if not merged.empty else validation

    analytics = build_override_analytics(merged)
    workbook_result = build_human_review_workbook(
        journal_df,
        scored_df=scored_df,
        lesson_df=lesson_df,
        analytics=analytics,
        output_path=diagnostics_dir / WORKBOOK_FILENAME,
    )
    gate = build_phase5w_release_gate(merged)

    validation.to_csv(diagnostics_dir / "phase5w01_human_review_validation.csv", index=False)
    analytics.to_csv(diagnostics_dir / "phase5w01_override_analytics.csv", index=False)
    merged.to_csv(diagnostics_dir / "phase5w01_human_review_merged_frame.csv", index=False)
    gate.to_csv(diagnostics_dir / "phase5w01_release_gate.csv", index=False)

    top_reasons = []
    if analytics.iloc[0].get("override_reasons_by_count"):
        top_reasons = str(analytics.iloc[0]["override_reasons_by_count"]).split(";")[:5]

    return {
        "workbook_generated": bool(workbook_result["workbook_generated"]),
        "workbook_path": workbook_result.get("workbook_path"),
        "review_rows": int(len(merged)),
        "completed_reviews": int(analytics.iloc[0]["completed_reviews"]),
        "pending_reviews": int(analytics.iloc[0]["pending_reviews"]),
        "review_completion_rate": float(analytics.iloc[0]["review_completion_rate"]),
        "override_count": int(analytics.iloc[0]["override_count"]),
        "accepted_brain_count": int(analytics.iloc[0]["accepted_brain_count"]),
        "accepted_governed_count": int(analytics.iloc[0]["accepted_governed_count"]),
        "average_human_confidence": float(analytics.iloc[0]["average_human_confidence"]),
        "top_override_reasons": top_reasons,
        "data_quality_blocks": int(analytics.iloc[0]["data_quality_blocks"]),
        "long_tail_overrides": int(analytics.iloc[0]["long_tail_overrides"]),
        "release_recommendation": "NO_RELEASE",
        "primary_blocker": "model_bias_dangerously_negative",
    }


def run_phase5w01_human_review_capture(*, diagnostics_dir: Path = DEFAULT_DIAGNOSTICS_DIR) -> dict[str, Any]:
    return write_phase5w_diagnostics(diagnostics_dir=diagnostics_dir)


def _action_difference_type(row: pd.Series) -> str:
    decision = str(row.get("human_buyer_decision", "")).strip()
    brain = str(row.get("brain_validated_action_label", "")).strip()
    governed = str(row.get("final_governed_action_label", "")).strip()
    if not decision:
        return "PENDING"
    if decision == "BUY_AS_BRAIN_SUGGESTED":
        return "ACCEPT_BRAIN"
    if decision == "BUY_AS_GOVERNED_SUGGESTED":
        return "ACCEPT_GOVERNED"
    if decision == "BUY_DIFFERENT_QUANTITY":
        return "QUANTITY_DIFFERENCE"
    if decision == "BLOCKED_DATA_QUALITY":
        return "DATA_QUALITY_BLOCK"
    if decision == "SUPPLIER_UNAVAILABLE":
        return "SUPPLIER_CONSTRAINT"
    if decision == "NO_BUY_RUN_DOWN":
        return "RUN_DOWN"
    if brain != governed:
        return "BRAIN_GOVERNED_MISMATCH"
    return "OTHER_HUMAN_PATH"


def _learning_theme_for_row(row: pd.Series) -> str:
    decision = str(row.get("human_buyer_decision", "")).strip()
    reason = str(row.get("human_override_reason", "")).strip()
    if not decision:
        return ""
    if decision == "BUY_AS_BRAIN_SUGGESTED":
        return "BUYER_ACCEPTS_BRAIN"
    if decision == "BUY_AS_GOVERNED_SUGGESTED":
        return "BUYER_ACCEPTS_GOVERNED"
    if reason == "BRAIN_TOO_AGGRESSIVE":
        return "BRAIN_TOO_AGGRESSIVE"
    if reason == "BRAIN_TOO_CONSERVATIVE":
        return "BRAIN_TOO_CONSERVATIVE"
    if reason == "GOVERNANCE_TOO_AGGRESSIVE":
        return "GOVERNANCE_TOO_AGGRESSIVE"
    if reason == "GOVERNANCE_TOO_CONSERVATIVE":
        return "GOVERNANCE_TOO_CONSERVATIVE"
    if reason in {"LONG_TAIL_BASKET_PROTECTION", "RANGE_TRUST_PROTECTION"}:
        return "LONG_TAIL_BASKET_PROTECTION"
    if decision == "SUPPLIER_UNAVAILABLE" or reason == "SUPPLIER_CONSTRAINT":
        return "SUPPLIER_CONSTRAINT"
    if decision == "BLOCKED_DATA_QUALITY" or reason == "LOW_CONFIDENCE_STOCK_DATA":
        return "DATA_QUALITY_BLOCK"
    if reason == "KNOWN_LOCAL_DEMAND":
        return "KNOWN_LOCAL_DEMAND"
    if str(row.get("human_overrides_both_flag", "")) == "YES":
        return "BUYER_OVERRIDES_BOTH"
    return "BUYER_OVERRIDES_BOTH"


def build_decision_quality_scorecard(merged_frame: pd.DataFrame) -> pd.DataFrame:
    """Build decision quality scorecard from merged human review frame."""
    base = build_override_analytics(merged_frame)
    if merged_frame.empty:
        return base

    status = merged_frame.get("human_review_status", pd.Series("PENDING", index=merged_frame.index)).astype(str)
    complete = status.eq("COMPLETE")
    conf = _numeric(merged_frame.get("human_confidence_score", pd.Series(0, index=merged_frame.index)))
    override_flag = merged_frame.get("human_override_flag", pd.Series("", index=merged_frame.index)).astype(str).str.upper().eq("YES")
    override_reason = merged_frame.get("human_override_reason", pd.Series("", index=merged_frame.index)).astype(str)
    decision = merged_frame.get("human_buyer_decision", pd.Series("", index=merged_frame.index)).astype(str)

    row = base.iloc[0].to_dict()
    row.update({
        "total_review_rows": int(len(merged_frame)),
        "high_confidence_override_count": int((complete & override_flag & conf.ge(70)).sum()),
        "low_confidence_decision_count": int((complete & conf.lt(50)).sum()),
        "long_tail_basket_protection_decisions": int(
            override_reason.isin({"LONG_TAIL_BASKET_PROTECTION", "RANGE_TRUST_PROTECTION"}).sum()
            + (
                decision.eq("BUY_DIFFERENT_QUANTITY")
                & merged_frame.get("long_tail_sku_flag", pd.Series("NO", index=merged_frame.index)).astype(str).eq("YES")
            ).sum()
        ),
        "governance_too_conservative_decisions": int(override_reason.eq("GOVERNANCE_TOO_CONSERVATIVE").sum()),
        "brain_too_aggressive_decisions": int(override_reason.eq("BRAIN_TOO_AGGRESSIVE").sum()),
        "brain_too_conservative_decisions": int(override_reason.eq("BRAIN_TOO_CONSERVATIVE").sum()),
    })
    return pd.DataFrame([row])


def build_human_decision_quality_rows(merged_frame: pd.DataFrame) -> pd.DataFrame:
    """Build row-level human decision quality output for completed reviews."""
    if merged_frame.empty:
        return pd.DataFrame()

    complete = merged_frame.get("human_review_status", pd.Series("", index=merged_frame.index)).astype(str).eq("COMPLETE")
    rows = merged_frame.loc[complete].copy()
    if rows.empty:
        return pd.DataFrame(columns=[
            "shadow_run_id", "store_number", "promotion_id", "sku_number",
            "brain_action", "governed_action", "human_action", "human_order_units",
            "human_confidence", "human_override_reason", "human_feedback_signal",
            "human_feedback_weight", "action_difference_type", "learning_note",
        ])

    out = pd.DataFrame({
        "shadow_run_id": rows.get("shadow_run_id", ""),
        "store_number": rows.get("store_number", ""),
        "promotion_id": rows.get("promotion_id", ""),
        "sku_number": rows.get("sku_number", ""),
        "brain_action": rows.get("brain_validated_action_label", ""),
        "governed_action": rows.get("final_governed_action_label", ""),
        "human_action": rows.get("human_buyer_decision", ""),
        "human_order_units": rows.get("human_order_units", ""),
        "human_confidence": rows.get("human_confidence_score", ""),
        "human_override_reason": rows.get("human_override_reason", ""),
        "human_feedback_signal": rows.get("human_feedback_signal", ""),
        "human_feedback_weight": rows.get("human_feedback_weight", ""),
        "action_difference_type": rows.apply(_action_difference_type, axis=1),
        "learning_note": rows.get("human_feedback_learning_note", ""),
    })
    return _round_frame(out)


def build_buyer_learning_pack(merged_frame: pd.DataFrame) -> pd.DataFrame:
    """Summarise human review into practical buyer learning themes."""
    if merged_frame.empty:
        return pd.DataFrame(columns=[
            "learning_theme", "row_count", "example_skus", "buyer_decision_pattern",
            "brain_learning_implication", "governance_learning_implication",
            "data_quality_implication", "recommended_next_action",
        ])

    complete = merged_frame.get("human_review_status", pd.Series("", index=merged_frame.index)).astype(str).eq("COMPLETE")
    reviewed = merged_frame.loc[complete].copy()
    if reviewed.empty:
        return pd.DataFrame([{
            "learning_theme": "PENDING_REVIEW",
            "row_count": int(len(merged_frame)),
            "example_skus": "",
            "buyer_decision_pattern": "No completed reviews yet",
            "brain_learning_implication": "MONITOR",
            "governance_learning_implication": "MONITOR",
            "data_quality_implication": "COMPLETE_FILLED_REVIEW_FILE",
            "recommended_next_action": "EXPORT_FILLED_REVIEW_CSV",
        }])

    reviewed = reviewed.copy()
    reviewed["_learning_theme"] = reviewed.apply(_learning_theme_for_row, axis=1)
    pack_rows = []
    theme_implications = {
        "BUYER_ACCEPTS_BRAIN": ("REINFORCE_BRAIN_WITH_REVIEW", "MONITOR", "NONE", "COMPARE_TO_ACTUAL_OUTCOMES"),
        "BUYER_ACCEPTS_GOVERNED": ("MONITOR", "REINFORCE_GOVERNANCE", "NONE", "COMPARE_TO_ACTUAL_OUTCOMES"),
        "BUYER_OVERRIDES_BOTH": ("REVIEW_ACTION_CLASSIFIER", "REVIEW_THRESHOLDS", "NONE", "AWAIT_OUTCOME_SCORING"),
        "BRAIN_TOO_AGGRESSIVE": ("REDUCE_BRAIN_AGGRESSION", "MONITOR", "NONE", "REVIEW_BRAIN_SEGMENTS"),
        "BRAIN_TOO_CONSERVATIVE": ("REVIEW_UNDERSTOCK_SEGMENTS", "MONITOR", "NONE", "REVIEW_BRAIN_SEGMENTS"),
        "GOVERNANCE_TOO_AGGRESSIVE": ("MONITOR", "REVIEW_FOR_TIGHTENING", "NONE", "HUMAN_APPROVAL_REQUIRED"),
        "GOVERNANCE_TOO_CONSERVATIVE": ("MONITOR", "REVIEW_FOR_RELAXATION", "NONE", "HUMAN_APPROVAL_REQUIRED"),
        "LONG_TAIL_BASKET_PROTECTION": ("REINFORCE_LONG_TAIL_FEATURES", "MAINTAIN_PROTECTION", "MAINTAIN_BASKET_EVIDENCE", "REINFORCE_BASKET_TRUST"),
        "SUPPLIER_CONSTRAINT": ("NO_BRAIN_FAILURE", "MONITOR", "REVIEW_SUPPLIER_DATA", "TRACK_SUPPLIER_RELIABILITY"),
        "DATA_QUALITY_BLOCK": ("NO_DEMAND_BRAIN_UPDATE", "BLOCK_CHANGE", "REPAIR_SOURCE_QUALITY", "FIX_DATA_FIRST"),
        "KNOWN_LOCAL_DEMAND": ("REVIEW_LOCAL_DEMAND_FEATURES", "MONITOR", "NONE", "CAPTURE_LOCAL_SIGNALS"),
    }

    for theme, grp in reviewed.groupby("_learning_theme", dropna=False):
        if not str(theme):
            continue
        brain_i, gov_i, dq_i, action = theme_implications.get(theme, ("MONITOR", "MONITOR", "NONE", "MONITOR"))
        examples = grp["sku_number"].astype(str).head(5).tolist()
        pack_rows.append({
            "learning_theme": theme,
            "row_count": int(len(grp)),
            "example_skus": ";".join(examples),
            "buyer_decision_pattern": ";".join(grp["human_buyer_decision"].astype(str).value_counts().head(3).index.tolist()),
            "brain_learning_implication": brain_i,
            "governance_learning_implication": gov_i,
            "data_quality_implication": dq_i,
            "recommended_next_action": action,
        })

    pack = pd.DataFrame(pack_rows).sort_values("row_count", ascending=False)
    for theme in BUYER_LEARNING_THEMES:
        if theme not in set(pack.get("learning_theme", pd.Series(dtype=str)).astype(str)):
            continue
    return pack if not pack.empty else pd.DataFrame([{
        "learning_theme": "PENDING_REVIEW",
        "row_count": 0,
        "example_skus": "",
        "buyer_decision_pattern": "No themed reviews",
        "brain_learning_implication": "MONITOR",
        "governance_learning_implication": "MONITOR",
        "data_quality_implication": "NONE",
        "recommended_next_action": "MONITOR",
    }])


def build_review_status_workbook(
    merged_frame: pd.DataFrame,
    *,
    scorecard: pd.DataFrame,
    learning_pack: pd.DataFrame,
    analytics: pd.DataFrame,
    output_path: Path,
) -> bool:
    """Write refreshed buyer review status workbook."""
    if not _excel_supported():
        return False

    status = merged_frame.get("human_review_status", pd.Series("PENDING", index=merged_frame.index)).astype(str)
    complete = status.eq("COMPLETE")
    pending = status.eq("PENDING")
    dq_blocks = merged_frame.loc[
        merged_frame.get("human_buyer_decision", pd.Series("", index=merged_frame.index)).astype(str).eq("BLOCKED_DATA_QUALITY")
    ].copy()

    identity_cols = [c for c in ("shadow_run_id", "store_number", "promotion_id", "sku_number", "sku_description", "human_buyer_decision", "human_review_status") if c in merged_frame.columns]
    sheets = {
        "Review_Status": _round_frame(merged_frame[identity_cols + [c for c in VALIDATION_COLUMNS + TRAINING_SIGNAL_COLUMNS if c in merged_frame.columns]]),
        "Completed_Reviews": _round_frame(merged_frame.loc[complete]),
        "Pending_Reviews": _round_frame(merged_frame.loc[pending]),
        "Override_Analytics": analytics,
        "Buyer_Learning_Pack": learning_pack,
        "Data_Quality_Blocks": _round_frame(dq_blocks) if not dq_blocks.empty else merged_frame.head(0),
        "Instructions": build_instructions_sheet(),
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, sheet_df in sheets.items():
            sheet_df.to_excel(writer, sheet_name=name, index=False)
    _style_workbook(output_path, fillable_columns=[])
    return output_path.exists()


def build_phase5x_release_gate(merged_frame: pd.DataFrame, *, filled_file_found: bool) -> pd.DataFrame:
    return pd.DataFrame([{
        "recommendation": "NO_RELEASE",
        "shadow_recommendation": "SHADOW_TOP_100_REVIEW",
        "filled_review_file_found": "YES" if filled_file_found else "NO",
        "review_rows": int(len(merged_frame)),
        "auto_order_created": "NO",
        "governed_actions_overwritten": "NO",
        "production_predictions_overwritten": "NO",
        "primary_blocker": "model_bias_dangerously_negative",
        "reason": "Filled human review learning is internal only; customer release not earned.",
    }])


def write_phase5x_diagnostics(
    *,
    journal_df: pd.DataFrame | None = None,
    filled_df: pd.DataFrame | None = None,
    filled_path: Path | None = None,
    scored_df: pd.DataFrame | None = None,
    lesson_df: pd.DataFrame | None = None,
    diagnostics_dir: Path = PHASE5X_DIAGNOSTICS_DIR,
) -> dict[str, Any]:
    """Ingest filled human review and write decision quality diagnostics."""
    diagnostics_dir.mkdir(parents=True, exist_ok=True)

    if journal_df is None:
        if not PHASE5T_JOURNAL_PATH.exists():
            raise FileNotFoundError(f"Shadow journal not found: {PHASE5T_JOURNAL_PATH}")
        journal_df = pd.read_csv(PHASE5T_JOURNAL_PATH)
    if scored_df is None and PHASE5U_SCORED_PATH.exists():
        scored_df = pd.read_csv(PHASE5U_SCORED_PATH)
    if lesson_df is None and PHASE5V_WEIGHTED_PATH.exists():
        lesson_df = pd.read_csv(PHASE5V_WEIGHTED_PATH)

    filled_file_found = False
    if filled_df is None:
        filled_df, filled_file_found = load_filled_human_review_file(filled_path or FILLED_REVIEW_PATH)
    else:
        filled_file_found = True

    validation = validate_filled_human_review_file(filled_df)
    merged = merge_human_review_feedback(journal_df, filled_df, scored_df=scored_df, lesson_df=lesson_df)
    if validation.empty:
        validation = merged[[c for c in list(MERGE_KEY_COLUMNS) + list(FILLABLE_COLUMNS) + list(VALIDATION_COLUMNS) if c in merged.columns]].copy()

    scorecard = build_decision_quality_scorecard(merged)
    quality_rows = build_human_decision_quality_rows(merged)
    learning_pack = build_buyer_learning_pack(merged)
    analytics = build_override_analytics(merged)
    gate = build_phase5x_release_gate(merged, filled_file_found=filled_file_found)

    validation.to_csv(diagnostics_dir / "phase5x01_filled_review_validation.csv", index=False)
    scorecard.to_csv(diagnostics_dir / "phase5x01_decision_quality_scorecard.csv", index=False)
    quality_rows.to_csv(diagnostics_dir / "phase5x01_human_decision_quality_rows.csv", index=False)
    learning_pack.to_csv(diagnostics_dir / "phase5x01_buyer_learning_pack.csv", index=False)
    gate.to_csv(diagnostics_dir / "phase5x01_release_gate.csv", index=False)
    status_workbook_generated = build_review_status_workbook(
        merged,
        scorecard=scorecard,
        learning_pack=learning_pack,
        analytics=analytics,
        output_path=diagnostics_dir / STATUS_WORKBOOK_FILENAME,
    )

    sc = scorecard.iloc[0] if not scorecard.empty else {}
    top_themes = learning_pack.sort_values("row_count", ascending=False)["learning_theme"].head(5).tolist() if not learning_pack.empty else []

    return {
        "filled_review_file_found": filled_file_found,
        "status_workbook_generated": status_workbook_generated,
        "review_rows": int(len(merged)),
        "completed_reviews": int(sc.get("completed_reviews", 0)),
        "pending_reviews": int(sc.get("pending_reviews", len(merged))),
        "review_completion_rate": float(sc.get("review_completion_rate", 0.0)),
        "accepted_brain_count": int(sc.get("accepted_brain_count", 0)),
        "accepted_governed_count": int(sc.get("accepted_governed_count", 0)),
        "override_count": int(sc.get("override_count", 0)),
        "high_confidence_override_count": int(sc.get("high_confidence_override_count", 0)),
        "long_tail_protection_decision_count": int(sc.get("long_tail_basket_protection_decisions", 0)),
        "data_quality_block_count": int(sc.get("blocked_data_quality_count", 0)),
        "average_human_confidence": float(sc.get("average_human_confidence", 0.0)),
        "top_buyer_learning_themes": top_themes,
        "release_recommendation": "NO_RELEASE",
        "primary_blocker": "model_bias_dangerously_negative",
    }


def run_phase5x01_filled_human_review_learning(*, diagnostics_dir: Path = PHASE5X_DIAGNOSTICS_DIR) -> dict[str, Any]:
    return write_phase5x_diagnostics(diagnostics_dir=diagnostics_dir)
