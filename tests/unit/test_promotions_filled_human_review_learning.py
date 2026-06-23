from __future__ import annotations

from pathlib import Path
import sys
import tempfile
import unittest

import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.append(str(REPO_ROOT / "src"))

from models.promotions.promo_human_review_capture import (  # noqa: E402
    MERGE_KEY_COLUMNS,
    build_buyer_learning_pack,
    build_decision_quality_scorecard,
    load_filled_human_review_file,
    merge_human_review_feedback,
    validate_filled_human_review_file,
    write_phase5x_diagnostics,
)
from models.promotions.promo_shadow_observation_journal import build_shadow_observation_journal  # noqa: E402
from tests.unit.test_promotions_promo_shadow_candidate_selection import _eligible_row  # noqa: E402


def _journal(n: int = 5) -> pd.DataFrame:
    from models.promotions.promo_shadow_candidate_selection import build_shadow_candidate_selection_frame

    rows = [_eligible_row(sku_number=str(100 + i), promotion_id=f"P{i}", economic_net_value_score=float(i + 1)) for i in range(n)]
    return build_shadow_observation_journal(
        build_shadow_candidate_selection_frame(pd.DataFrame(rows)),
        config={"shadow_run_id": "phase5x-test"},
    )


def _keys_row(journal: pd.DataFrame, idx: int = 0) -> dict:
    row = journal.iloc[idx]
    return {k: row[k] for k in MERGE_KEY_COLUMNS}


def _valid_human_row(journal: pd.DataFrame, idx: int = 0, **overrides: str) -> dict:
    base = {
        **_keys_row(journal, idx),
        "human_buyer_decision": "BUY_AS_BRAIN_SUGGESTED",
        "human_order_units": "2",
        "human_decision_reason": "Brain path preferred",
        "human_override_flag": "YES",
        "human_override_reason": "BRAIN_TOO_CONSERVATIVE",
        "human_confidence_score": "85",
        "human_reviewer": "buyer1",
        "human_reviewed_at": "2026-06-22",
        "human_notes": "",
    }
    base.update(overrides)
    return base


class TestFilledHumanReviewLearning(unittest.TestCase):
    def test_missing_filled_file_runs_gracefully(self) -> None:
        journal = _journal(4)
        with tempfile.TemporaryDirectory() as tmp:
            diag = Path(tmp)
            result = write_phase5x_diagnostics(journal_df=journal, filled_path=diag / "missing.csv", diagnostics_dir=diag)
            self.assertFalse(result["filled_review_file_found"])
            self.assertEqual(result["pending_reviews"], 4)
            self.assertEqual(result["release_recommendation"], "NO_RELEASE")
            self.assertTrue((diag / "phase5x01_decision_quality_scorecard.csv").exists())

    def test_valid_filled_review_merges(self) -> None:
        journal = _journal(3)
        human = pd.DataFrame([_valid_human_row(journal, 0), _valid_human_row(journal, 1, human_buyer_decision="BUY_AS_GOVERNED_SUGGESTED", human_override_flag="NO")])
        merged = merge_human_review_feedback(journal, human)
        self.assertEqual(int(merged["human_review_status"].eq("COMPLETE").sum()), 2)
        self.assertEqual(int(merged["human_review_status"].eq("PENDING").sum()), 1)

    def test_duplicate_keys_flagged(self) -> None:
        journal = _journal(1)
        human = pd.DataFrame([_valid_human_row(journal), _valid_human_row(journal, human_reviewer="buyer2")])
        validated = validate_filled_human_review_file(human)
        self.assertTrue(validated["human_review_validation_error"].astype(str).str.contains("duplicate").any())

    def test_invalid_decision_flagged(self) -> None:
        journal = _journal(1)
        human = pd.DataFrame([_valid_human_row(journal, human_buyer_decision="BAD_DECISION")])
        validated = validate_filled_human_review_file(human)
        self.assertEqual(validated.iloc[0]["human_review_status"], "INVALID")

    def test_pending_rows_preserved(self) -> None:
        journal = _journal(3)
        human = journal[list(MERGE_KEY_COLUMNS)].copy()
        human["human_buyer_decision"] = ""
        validated = validate_filled_human_review_file(human)
        self.assertEqual(int(validated["human_review_status"].eq("PENDING").sum()), 3)

    def test_accepted_brain_flag(self) -> None:
        journal = _journal(1)
        merged = merge_human_review_feedback(journal, pd.DataFrame([_valid_human_row(journal)]))
        self.assertEqual(merged.iloc[0]["human_agrees_with_brain_flag"], "YES")

    def test_accepted_governed_flag(self) -> None:
        journal = _journal(1)
        row = _valid_human_row(journal, human_buyer_decision="BUY_AS_GOVERNED_SUGGESTED", human_override_flag="NO", human_override_reason="")
        merged = merge_human_review_feedback(journal, pd.DataFrame([row]))
        self.assertEqual(merged.iloc[0]["human_agrees_with_governed_flag"], "YES")

    def test_override_both_flag(self) -> None:
        journal = _journal(1)
        row = _valid_human_row(journal, human_buyer_decision="BUY_DIFFERENT_QUANTITY", human_override_flag="YES", human_override_reason="OTHER")
        merged = merge_human_review_feedback(journal, pd.DataFrame([row]))
        self.assertEqual(merged.iloc[0]["human_overrides_both_flag"], "YES")

    def test_confidence_weighted_signal(self) -> None:
        journal = _journal(1)
        merged = merge_human_review_feedback(journal, pd.DataFrame([_valid_human_row(journal, human_confidence_score="80")]))
        self.assertGreater(float(merged.iloc[0]["human_confidence_weighted_signal"]), 0.0)

    def test_data_quality_block_not_demand_brain_training(self) -> None:
        journal = _journal(1)
        row = _valid_human_row(journal, human_buyer_decision="BLOCKED_DATA_QUALITY", human_override_reason="LOW_CONFIDENCE_STOCK_DATA")
        merged = merge_human_review_feedback(journal, pd.DataFrame([row]))
        self.assertEqual(merged.iloc[0]["human_feedback_training_ready_flag"], "NO")
        self.assertIn("data-quality", merged.iloc[0]["human_feedback_learning_note"].lower())

    def test_supplier_unavailable_not_brain_failure(self) -> None:
        journal = _journal(1)
        row = _valid_human_row(journal, human_buyer_decision="SUPPLIER_UNAVAILABLE", human_override_reason="SUPPLIER_CONSTRAINT")
        merged = merge_human_review_feedback(journal, pd.DataFrame([row]))
        self.assertIn("supplier", merged.iloc[0]["human_feedback_learning_note"].lower())

    def test_scorecard_reconciles_counts(self) -> None:
        journal = _journal(4)
        human = pd.DataFrame([_valid_human_row(journal, 0), _valid_human_row(journal, 1, human_buyer_decision="HOLD", human_order_units="0", human_override_flag="YES")])
        merged = merge_human_review_feedback(journal, human)
        scorecard = build_decision_quality_scorecard(merged)
        self.assertEqual(int(scorecard.iloc[0]["total_review_rows"]), 4)
        self.assertEqual(int(scorecard.iloc[0]["completed_reviews"]), 2)
        self.assertEqual(int(scorecard.iloc[0]["pending_reviews"]), 2)

    def test_refreshed_workbook_created(self) -> None:
        journal = _journal(3)
        with tempfile.TemporaryDirectory() as tmp:
            diag = Path(tmp)
            result = write_phase5x_diagnostics(journal_df=journal, diagnostics_dir=diag)
            self.assertTrue(result["status_workbook_generated"])
            from openpyxl import load_workbook

            wb = load_workbook(diag / "SHADOW_TOP_100_BUYER_REVIEW_STATUS.xlsx")
            for sheet in ("Review_Status", "Completed_Reviews", "Pending_Reviews", "Override_Analytics", "Buyer_Learning_Pack", "Data_Quality_Blocks", "Instructions"):
                self.assertIn(sheet, wb.sheetnames)

    def test_buyer_learning_pack_themes(self) -> None:
        journal = _journal(2)
        human = pd.DataFrame([
            _valid_human_row(journal, 0),
            _valid_human_row(journal, 1, human_buyer_decision="BUY_AS_GOVERNED_SUGGESTED", human_override_flag="NO", human_override_reason=""),
        ])
        merged = merge_human_review_feedback(journal, human)
        pack = build_buyer_learning_pack(merged)
        self.assertTrue(set(pack["learning_theme"]).issubset({
            "BUYER_ACCEPTS_BRAIN", "BUYER_ACCEPTS_GOVERNED", "BUYER_OVERRIDES_BOTH",
            "BRAIN_TOO_AGGRESSIVE", "BRAIN_TOO_CONSERVATIVE", "GOVERNANCE_TOO_AGGRESSIVE",
            "GOVERNANCE_TOO_CONSERVATIVE", "LONG_TAIL_BASKET_PROTECTION", "SUPPLIER_CONSTRAINT",
            "DATA_QUALITY_BLOCK", "KNOWN_LOCAL_DEMAND", "PENDING_REVIEW",
        }))

    def test_governed_actions_not_overwritten(self) -> None:
        journal = _journal(2)
        before = journal["final_governed_action_label"].copy()
        merged = merge_human_review_feedback(journal, pd.DataFrame([_valid_human_row(journal, 0)]))
        self.assertTrue((before == merged["final_governed_action_label"]).all())

    def test_no_order_file_created(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            diag = Path(tmp)
            write_phase5x_diagnostics(journal_df=_journal(3), diagnostics_dir=diag)
            self.assertEqual(list(diag.glob("*order*")), [])

    def test_load_filled_file_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            frame, found = load_filled_human_review_file(Path(tmp) / "nope.csv")
            self.assertFalse(found)
            self.assertIsNone(frame)


if __name__ == "__main__":
    unittest.main()
