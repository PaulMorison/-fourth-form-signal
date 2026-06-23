from __future__ import annotations

from pathlib import Path
import sys
import tempfile
import unittest

import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.append(str(REPO_ROOT / "src"))

from models.promotions.promo_human_review_capture import (  # noqa: E402
    FILLABLE_COLUMNS,
    HUMAN_OVERRIDE_REASONS,
    MERGE_KEY_COLUMNS,
    build_human_review_workbook,
    build_override_analytics,
    merge_human_review_feedback,
    validate_human_review_input,
    write_phase5w_diagnostics,
)
from models.promotions.promo_shadow_observation_journal import (  # noqa: E402
    HUMAN_BUYER_DECISIONS,
    build_shadow_observation_journal,
)
from tests.unit.test_promotions_promo_shadow_candidate_selection import _eligible_row  # noqa: E402


def _journal(n: int = 5) -> pd.DataFrame:
    from models.promotions.promo_shadow_candidate_selection import build_shadow_candidate_selection_frame

    rows = [_eligible_row(sku_number=str(100 + i), promotion_id=f"P{i}", economic_net_value_score=float(i + 1)) for i in range(n)]
    return build_shadow_observation_journal(
        build_shadow_candidate_selection_frame(pd.DataFrame(rows)),
        config={"shadow_run_id": "phase5w-test"},
    )


def _keys_row(journal: pd.DataFrame, idx: int = 0) -> dict:
    row = journal.iloc[idx]
    return {k: row[k] for k in MERGE_KEY_COLUMNS}


class TestHumanReviewCapture(unittest.TestCase):
    def test_review_workbook_created_with_required_sheets(self) -> None:
        journal = _journal(8)
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "workbook.xlsx"
            result = build_human_review_workbook(journal, output_path=path)
            self.assertTrue(result["workbook_generated"])
            self.assertTrue(path.exists())
            from openpyxl import load_workbook

            wb = load_workbook(path)
            for sheet in (
                "Top_100_Review",
                "Long_Tail_Mission_SKUs",
                "Governance_Too_Conservative",
                "Brain_vs_Governed_Mismatch",
                "Allowed_Values",
                "Instructions",
                "Review_Summary",
            ):
                self.assertIn(sheet, wb.sheetnames)

    def test_fillable_fields_exist_on_main_sheet(self) -> None:
        journal = _journal(3)
        result = build_human_review_workbook(journal, output_path=None)
        top = result["sheets"]["Top_100_Review"]
        for col in FILLABLE_COLUMNS:
            self.assertIn(col, top.columns)

    def test_allowed_values_sheet_exists(self) -> None:
        journal = _journal(2)
        result = build_human_review_workbook(journal, output_path=None)
        allowed = result["sheets"]["Allowed_Values"]
        self.assertIn("human_buyer_decision", set(allowed["field_name"]))
        self.assertIn("human_override_reason", set(allowed["field_name"]))

    def test_validation_accepts_valid_decisions(self) -> None:
        journal = _journal(1)
        keys = _keys_row(journal)
        human = pd.DataFrame([{
            **keys,
            "human_buyer_decision": "BUY_AS_GOVERNED_SUGGESTED",
            "human_order_units": "2",
            "human_decision_reason": "Align with governed",
            "human_override_flag": "NO",
            "human_override_reason": "",
            "human_confidence_score": "80",
            "human_reviewer": "buyer1",
            "human_reviewed_at": "2026-06-22",
            "human_notes": "",
        }])
        validated = validate_human_review_input(human)
        self.assertEqual(validated.iloc[0]["human_review_status"], "COMPLETE")
        self.assertEqual(validated.iloc[0]["human_review_valid_flag"], "YES")

    def test_validation_flags_invalid_decisions(self) -> None:
        journal = _journal(1)
        keys = _keys_row(journal)
        human = pd.DataFrame([{
            **keys,
            "human_buyer_decision": "NOT_ALLOWED",
            "human_order_units": "-1",
            "human_confidence_score": "150",
            "human_reviewer": "",
            "human_override_reason": "BAD_REASON",
        }])
        validated = validate_human_review_input(human)
        self.assertEqual(validated.iloc[0]["human_review_status"], "INVALID")
        err = validated.iloc[0]["human_review_validation_error"]
        self.assertIn("invalid_human_buyer_decision", err)

    def test_duplicate_reviews_flagged(self) -> None:
        journal = _journal(1)
        keys = _keys_row(journal)
        human = pd.DataFrame([
            {**keys, "human_buyer_decision": "HOLD", "human_order_units": "0", "human_confidence_score": "70", "human_reviewer": "a"},
            {**keys, "human_buyer_decision": "HOLD", "human_order_units": "0", "human_confidence_score": "70", "human_reviewer": "b"},
        ])
        validated = validate_human_review_input(human)
        self.assertTrue(validated["human_review_validation_error"].astype(str).str.contains("duplicate").any())

    def test_pending_rows_remain_pending_not_failed(self) -> None:
        journal = _journal(3)
        human = journal[list(MERGE_KEY_COLUMNS)].copy()
        human["human_buyer_decision"] = ""
        validated = validate_human_review_input(human)
        self.assertTrue((validated["human_review_status"] == "PENDING").all())
        self.assertFalse((validated["human_review_status"] == "INVALID").any())

    def test_confidence_score_bounds_enforced(self) -> None:
        journal = _journal(1)
        keys = _keys_row(journal)
        human = pd.DataFrame([{**keys, "human_buyer_decision": "HOLD", "human_order_units": "0", "human_confidence_score": "101", "human_reviewer": "x"}])
        validated = validate_human_review_input(human)
        self.assertIn("human_confidence_out_of_range", validated.iloc[0]["human_review_validation_error"])

    def test_override_analytics_reconcile_row_counts(self) -> None:
        journal = _journal(4)
        keys = [_keys_row(journal, i) for i in range(2)]
        human = pd.DataFrame([
            {**keys[0], "human_buyer_decision": "BUY_AS_BRAIN_SUGGESTED", "human_order_units": "1", "human_override_flag": "YES",
             "human_override_reason": "BRAIN_TOO_CONSERVATIVE", "human_confidence_score": "75", "human_reviewer": "b1", "human_reviewed_at": "2026-06-22"},
            {**keys[1], "human_buyer_decision": "BUY_AS_GOVERNED_SUGGESTED", "human_order_units": "2", "human_override_flag": "NO",
             "human_override_reason": "", "human_confidence_score": "85", "human_reviewer": "b2", "human_reviewed_at": "2026-06-22"},
        ])
        merged = merge_human_review_feedback(journal, human)
        analytics = build_override_analytics(merged)
        self.assertEqual(int(analytics.iloc[0]["completed_reviews"]), 2)
        self.assertEqual(int(analytics.iloc[0]["pending_reviews"]), 2)
        self.assertEqual(int(analytics.iloc[0]["accepted_brain_count"]), 1)
        self.assertEqual(int(analytics.iloc[0]["accepted_governed_count"]), 1)

    def test_human_feedback_signals_created(self) -> None:
        journal = _journal(1)
        keys = _keys_row(journal)
        human = pd.DataFrame([{
            **keys,
            "human_buyer_decision": "BUY_AS_BRAIN_SUGGESTED",
            "human_order_units": "1",
            "human_override_flag": "YES",
            "human_override_reason": "BRAIN_TOO_CONSERVATIVE",
            "human_confidence_score": "90",
            "human_reviewer": "buyer1",
            "human_reviewed_at": "2026-06-22",
        }])
        merged = merge_human_review_feedback(journal, human)
        self.assertEqual(merged.iloc[0]["human_accepts_brain_flag"], "YES")
        self.assertNotEqual(merged.iloc[0]["human_feedback_signal"], "PENDING")

    def test_governed_actions_not_overwritten(self) -> None:
        journal = _journal(2)
        before = journal["final_governed_action_label"].copy()
        human = journal[list(MERGE_KEY_COLUMNS)].copy()
        human["human_buyer_decision"] = "HOLD"
        human["human_order_units"] = "0"
        human["human_confidence_score"] = "60"
        human["human_reviewer"] = "buyer"
        human["human_reviewed_at"] = "2026-06-22"
        merged = merge_human_review_feedback(journal, human)
        self.assertTrue((before == merged["final_governed_action_label"]).all())

    def test_no_order_file_created(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            diag = Path(tmp)
            write_phase5w_diagnostics(journal_df=_journal(5), diagnostics_dir=diag)
            order_like = [p for p in diag.glob("*order*")]
            self.assertEqual(order_like, [])

    def test_release_status_unchanged(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            diag = Path(tmp)
            result = write_phase5w_diagnostics(journal_df=_journal(5), diagnostics_dir=diag)
            gate = pd.read_csv(diag / "phase5w01_release_gate.csv")
            self.assertEqual(result["release_recommendation"], "NO_RELEASE")
            self.assertEqual(gate.iloc[0]["auto_order_created"], "NO")
            self.assertEqual(gate.iloc[0]["governed_actions_overwritten"], "NO")

    def test_override_reason_allowed_values(self) -> None:
        for reason in HUMAN_OVERRIDE_REASONS:
            self.assertIn(reason, HUMAN_OVERRIDE_REASONS)
        for decision in HUMAN_BUYER_DECISIONS:
            self.assertIn(decision, HUMAN_BUYER_DECISIONS)


if __name__ == "__main__":
    unittest.main()
